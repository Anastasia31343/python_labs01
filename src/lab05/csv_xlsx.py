import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)
    
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    rows = []
    with csv_file.open('r', encoding='utf-8') as f:
        try:
            reader = csv.reader(f)
            rows = list(reader)
        except csv.Error as e:
            raise ValueError(f"Invalid CSV format: {e}")
    
    if not rows:
        raise ValueError("Empty CSV file")
    
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    
    for row_idx, row_data in enumerate(rows, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
    
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = max(max_length + 2, 8)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(xlsx_file)